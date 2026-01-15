import streamlit as st
import pandas as pd
from openpyxl.styles import Font, PatternFill
from io import BytesIO
import openpyxl

st.title("Extraction de données mensuelles")
st.write(
    "App pour l'extraction de données mensuelles."
)
# File uploader widget
uploaded_file = st.file_uploader("Chargez votre fichier de statistiques mensuelles", type=["xlsx"])

if uploaded_file is not None:
    population_data = pd.read_excel(uploaded_file, sheet_name = 'population')
    education_data = pd.read_excel(uploaded_file, sheet_name = 'education')

    ##datacleanup and processing population data
    #drop les lignes ne contenant aucune donnée
    population_data = population_data.dropna(subset = ['coa_admin1'])
    population_data = population_data.drop(columns=['asylum','m_18-59','f_18-59'])
    for col in population_data.columns:
        if population_data[col].dtype == 'float64':
            population_data[col] = population_data[col].fillna(0).astype(int) 

    population_data['m_18-59'] = population_data[['m_18-24', 'm_25-49','m_50-59']].sum(axis=1)
    population_data['f_18-59'] = population_data[['f_18-24', 'f_25-49','f_50-59']].sum(axis=1)
    female_age_cols = ["f_0-4", "f_5-11", "f_12-17", "f_18-24", "f_25-49", "f_50-59", "f_60", "f_NA"]

    # 3) Calculer la somme pour les lignes 'detailed' uniquement
    mask_detailed = population_data["aggregation_type"].str.lower().eq("detailed")

    # Somme des tranches d’âge féminines (NaN ignorés)
    female_sum = population_data.loc[mask_detailed, female_age_cols].sum(axis=1, skipna=True)

    # 4) Mettre à jour 'female' **seulement** pour les lignes 'detailed'
    population_data.loc[mask_detailed, "female"] = female_sum

    # (Optionnel) Si tu veux t’assurer que 'female' est un entier lorsque c’est possible :
    population_data["female"] = population_data["female"].round().astype("Int64")

    ##datacleanup and processing education data
    #drop les lignes ne contenant aucune donnée
    education_data = education_data.dropna(subset = ['coa_admin1']) 

    #drop la colonne asylum
    education_data = education_data.drop(columns=['asylum'])
    # remplacer les valeurs NaN par 0 et convertir les colonnes float en int
    for col in education_data.columns:
        if education_data[col].dtype == 'float64':
            education_data[col] = education_data[col].fillna(0).astype(int)

    region = population_data['coa_admin1'].unique()
    def insert_between(columns, left_col, target_col, right_col):
        """
        Place target_col entre left_col et right_col si left_col et target_col existent.
        Si right_col n'existe pas, on met target_col juste après left_col.
        """
        cols = list(columns)
        if target_col in cols:
                cols.remove(target_col)
        else:
            # si la colonne n'existe pas, rien à faire
            return cols

        if left_col in cols:
            insert_pos = cols.index(left_col) + 1
            cols.insert(insert_pos, target_col)
        else:
            # si left_col n'existe pas, on remet target_col à la fin pour ne pas le perdre
            cols.append(target_col)
            return cols

        # Si right_col existe et se retrouve avant target_col, on réajuste
        if right_col in cols:
            # s'assurer que l'ordre soit left_col, target_col, right_col
            # si right_col s'est retrouvé avant, on le laisse; l'essentiel est que target soit après left_col
            pass

        return cols

    cols = list(population_data.columns)
    cols = insert_between(cols, "m_50-59", "m_18-59", "m_60")
    cols = insert_between(cols, "f_50-59", "f_18-59", "f_60")

    population_data = population_data.reindex(columns=cols)
      
    
    
    
    # Create dropdown in Streamlit
    region_choisie = st.selectbox("Sélectionnez une région", region)
    st.write(f"vous avez sélectionné: {region_choisie}")
    population_data_region = population_data[population_data['coa_admin1'] == region_choisie]
    education_data_region = education_data[education_data['coa_admin1'] == region_choisie]
    
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        population_data_region.to_excel(writer, sheet_name="Population", index=False)
        education_data_region.to_excel(writer, sheet_name="Education", index=False)
    # Style headers & set column widths for each sheet
    for sheet_name in ["Population", "Education"]:
        ws = writer.sheets[sheet_name]

        # Bold white headers on blue background
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4F81BD")

        # Auto width (basic heuristic)
        for col in ws.columns:
            values = [str(c.value) if c.value is not None else "" for c in col]
            max_len = max((len(v) for v in values), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
  
    buffer.seek(0)
    st.download_button(
        "⬇️ Download Excel (2 sheets)",
        data=buffer,
        file_name=f"{region_choisie}_extracted.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)


